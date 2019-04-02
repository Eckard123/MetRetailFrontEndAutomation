import openpyxl
import openpyxl_templates
import openpyxl_utilities


class Inputs:

    Column_Headers_FELEMENTL1 = []
    Column_Headers_FBPELEMEPF = []
    ELEINTFPR = []
    ELERENFPR = []
    ELEINTVPR = []
    ELERENVPR = []
    ELEDCSFPM = []
    ELEIAWSFP = []
    ELEIUFSFP = []
    ELEIUFRSF = []
    ELERISKPR = []
    ELELREPRE = []
    ELEPRFPRE = []
    ELECLMPRE = []

    Concatenate = []
    workbook = None
    worksheet = None
    maximum_rows = None

    PBEIFFBAL = []
    PBERFFBAL = []
    PBEIVFBAL = []
    PBERVFBAL = []
    PBEDSFBAL = []
    PBEINTAWB = []
    PBEINUFSB = []
    PBEINUFLB = []
    PBERISKBL = []
    PBERSKCHR = []
    PBELREBAL = []
    PBEPMPBAL = []
    PBECEPBAL = []


# Read workbook
Inputs.workbook = openpyxl.load_workbook("C:\\Users\\EcBerry\\Desktop\\Copy of METPOLS   20190401.xlsx")
Inputs.worksheet = Inputs.workbook["FELEMENTL1"]
Inputs.maximum_rows = Inputs.worksheet.max_row
print(Inputs.worksheet)
print(Inputs.maximum_rows)

# Assign all columns in FELEMENTL1 sheet to Column_Headers_FELEMENTL1
for row in Inputs.worksheet.iter_rows(min_row=0, min_col=0, max_col=26, max_row=1):
    for cell in row:
        Inputs.Column_Headers_FELEMENTL1.append(cell.value)
        print(cell.value, end=" ")  # Print next to each other not below each other

# Assign all values in each element column to that column
print(end="\n")
for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FELEMENTL1.index('ELEINTFPR') + 1, max_col=Inputs.Column_Headers_FELEMENTL1.index('ELEINTFPR') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.ELEINTFPR.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FELEMENTL1.index('ELERENFPR') + 1, max_col=Inputs.Column_Headers_FELEMENTL1.index('ELERENFPR') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.ELERENFPR.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FELEMENTL1.index('ELEINTVPR') + 1, max_col=Inputs.Column_Headers_FELEMENTL1.index('ELEINTVPR') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.ELEINTVPR.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FELEMENTL1.index('ELERENVPR') + 1, max_col=Inputs.Column_Headers_FELEMENTL1.index('ELERENVPR') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.ELERENVPR.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FELEMENTL1.index('ELEDCSFPM') + 1, max_col=Inputs.Column_Headers_FELEMENTL1.index('ELEDCSFPM') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.ELEDCSFPM.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FELEMENTL1.index('ELEIAWSFP') + 1, max_col=Inputs.Column_Headers_FELEMENTL1.index('ELEIAWSFP') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.ELEIAWSFP.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FELEMENTL1.index('ELEIUFSFP') + 1, max_col=Inputs.Column_Headers_FELEMENTL1.index('ELEIUFSFP') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.ELEIUFSFP.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FELEMENTL1.index('ELEIUFRSF') + 1, max_col=Inputs.Column_Headers_FELEMENTL1.index('ELEIUFRSF') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.ELEIUFRSF.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FELEMENTL1.index('ELERISKPR') + 1, max_col=Inputs.Column_Headers_FELEMENTL1.index('ELERISKPR') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.ELERISKPR.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FELEMENTL1.index('ELELREPRE') + 1, max_col=Inputs.Column_Headers_FELEMENTL1.index('ELELREPRE') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.ELELREPRE.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FELEMENTL1.index('ELEPRFPRE') + 1, max_col=Inputs.Column_Headers_FELEMENTL1.index('ELEPRFPRE') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.ELEPRFPRE.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FELEMENTL1.index('ELECLMPRE') + 1, max_col=Inputs.Column_Headers_FELEMENTL1.index('ELECLMPRE') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.ELECLMPRE.append(cell.value)

# Now for FBPELEMEPF
Inputs.worksheet = Inputs.workbook["FBPELEMEPF"]
Inputs.maximum_rows = Inputs.worksheet.max_row

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=0, max_col=27, max_row=1):
    for cell in row:
        Inputs.Column_Headers_FBPELEMEPF.append(cell.value)
        print(cell.value, end=" ")  # Print next to each other not below each other

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FBPELEMEPF.index('PBEIFFBAL') + 1, max_col=Inputs.Column_Headers_FBPELEMEPF.index('PBEIFFBAL') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.PBEIFFBAL.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FBPELEMEPF.index('PBERFFBAL') + 1, max_col=Inputs.Column_Headers_FBPELEMEPF.index('PBERFFBAL') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.PBERFFBAL.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FBPELEMEPF.index('PBEIVFBAL') + 1, max_col=Inputs.Column_Headers_FBPELEMEPF.index('PBEIVFBAL') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.PBEIVFBAL.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FBPELEMEPF.index('PBERVFBAL') + 1, max_col=Inputs.Column_Headers_FBPELEMEPF.index('PBERVFBAL') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.PBERVFBAL.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FBPELEMEPF.index('PBEDSFBAL') + 1, max_col=Inputs.Column_Headers_FBPELEMEPF.index('PBEDSFBAL') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.PBEDSFBAL.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FBPELEMEPF.index('PBEINTAWB') + 1, max_col=Inputs.Column_Headers_FBPELEMEPF.index('PBEINTAWB') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.PBEINTAWB.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FBPELEMEPF.index('PBEINUFSB') + 1, max_col=Inputs.Column_Headers_FBPELEMEPF.index('PBEINUFSB') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.PBEINUFSB.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FBPELEMEPF.index('PBEINUFLB') + 1, max_col=Inputs.Column_Headers_FBPELEMEPF.index('PBEINUFLB') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.PBEINUFLB.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FBPELEMEPF.index('PBERISKBL') + 1, max_col=Inputs.Column_Headers_FBPELEMEPF.index('PBERISKBL') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.PBERISKBL.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FBPELEMEPF.index('PBERSKCHR') + 1, max_col=Inputs.Column_Headers_FBPELEMEPF.index('PBERSKCHR') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.PBERSKCHR.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FBPELEMEPF.index('PBELREBAL') + 1, max_col=Inputs.Column_Headers_FBPELEMEPF.index('PBELREBAL') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.PBELREBAL.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FBPELEMEPF.index('PBEPMPBAL') + 1, max_col=Inputs.Column_Headers_FBPELEMEPF.index('PBEPMPBAL') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.PBEPMPBAL.append(cell.value)

for row in Inputs.worksheet.iter_rows(min_row=0, min_col=Inputs.Column_Headers_FBPELEMEPF.index('PBECEPBAL') + 1, max_col=Inputs.Column_Headers_FBPELEMEPF.index('PBECEPBAL') + 1, max_row=Inputs.maximum_rows):
    for cell in row:
        Inputs.PBECEPBAL.append(cell.value)








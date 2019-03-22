import xlwings
import xlrd
import datetime
import math as mathtool
import openpyxl, openpyxl_templates, openpyxl_utilities
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.colors import GREEN

scenario_source = "C:\\Users\\EcBerry\\Desktop\\7005135.xlsx"
old_tool = "C:\\Users\\EcBerry\\Desktop\\Myriad info\\Metropolitain\\VP\\Peter & Danelle pricing tools\\pk Alteration tool.xlsb"
Todays_date = datetime.date.today()
print("Gesegnet bist du, wenn andere dich schmähen und dich verfolgen und allerlei Böses gegen mich falsch machen.")


def the_date_maker(SomeFunction):

    def newbuss_date(SomeDate):
        newbuss_year = int(SomeDate.year - 1)
        start_date = SomeDate.replace(day=1, month=1, year=newbuss_year)
        return start_date.strftime("%Y/%m/%d")

    def create_dob(currentdate, excel_age):
        dob_year = currentdate.year
        new_year = int(dob_year - excel_age)
        created_dob = currentdate.replace(day=1, month=1, year=new_year)
        print(created_dob.strftime("%Y/%m/%d"))
        return created_dob.strftime("%Y/%m/%d")
    return create_dob

@the_date_maker
def scenario_dob(date, ageofperson):
    return scenario_dob


class input:

    age = ""
    table_code_1 = ""
    table_code_2 = ""
    table_code_3 = ""
    table_code_4 = ""
    dob = None
    sum_assured = None
    fundopt = "STLV"
    growth_Percentage = None
    package_code = None
    scheme_code = None


class output:

    main_tot_before_VP = None
    main_tot_after_VP = None
    second_tot_before_VP = None
    second_tot_after_VP = None
    third_tot_before_VP = None
    third_tot_after_VP = None
    fourth_tot_before_VP = None
    fourth_tot_after_VP = None

    main_tot_elements_before_VP = None
    main_tot_elements_after_VP = None
    pudt_tot_elements_before_VP = None
    pudt_tot_elements_after_VP = None
    puds_tot_elements_before_VP = None
    puds_tot_elements_after_VP = None
    pure_o_tot_elements_before_VP = None
    pure_o_tot_elements_after_VP = None

    ProfitMarginPremBeforeVP_Main = None
    CleanRiskPremiumBeforeVP_Main = None
    initFixedPFeeBeforeVP_Main = None
    renewalFixedPFeeBeforeVP_Main = None
    claimsExpPremiumBeforeVP_Main = None
    renewalVariPFeeBeforeVP_Main = None
    DisribChannelSalesBeforeVP_Main = None
    InterAWSaleBeforeVP_Main = None
    InterUFSaleBeforeVP_Main = None
    InterUFLapseBeforeVP_Main = None
    LapseRiskOnExpBeforeVP_Main = None

    ProfitMarginPremAfterVP_Main = None
    CleanRiskPremiumAfterVP_Main = None
    initFixedPFeeAfterVP_Main = None
    renewalFixedPFeeAfterVP_Main = None
    claimsExpPremiumAfterVP_Main = None
    renewalVariPFeeAfterVP_Main = None
    DisribChannelSalesAfterVP_Main = None
    InterAWSaleAfterVP_Main = None
    InterUFSaleAfterVP_Main = None
    InterUFLapseAfterVP_Main = None
    LapseRiskOnExpAfterVP_Main = None

    ProfitMarginPremBeforeVP_Second = None
    CleanRiskPremiumBeforeVP_Second = None
    initFixedPFeeBeforeVP_Second = None
    renewalFixedPFeeBeforeVP_Second = None
    claimsExpPremiumBeforeVP_Second = None
    renewalVariPFeeBeforeVP_Second = None
    DisribChannelSalesBeforeVP_Second = None
    InterAWSaleBeforeVP_Second = None
    InterUFSaleBeforeVP_Second = None
    InterUFLapseBeforeVP_Second = None
    LapseRiskOnExpBeforeVP_Second = None

    ProfitMarginPremAfterVP_Second = None
    CleanRiskPremiumAfterVP_Second = None
    initFixedPFeeAfterVP_Second = None
    renewalFixedPFeeAfterVP_Second = None
    claimsExpPremiumAfterVP_Second = None
    renewalVariPFeeAfterVP_Second = None
    DisribChannelSalesAfterVP_Second = None
    InterAWSaleAfterVP_Second = None
    InterUFSaleAfterVP_Second = None
    InterUFLapseAfterVP_Second = None
    LapseRiskOnExpAfterVP_Second = None

    ProfitMarginPremBeforeVP_Third = None
    CleanRiskPremiumBeforeVP_Third = None
    initFixedPFeeBeforeVP_Third = None
    renewalFixedPFeeBeforeVP_Third = None
    claimsExpPremiumBeforeVP_Third = None
    renewalVariPFeeBeforeVP_Third = None
    DisribChannelSalesBeforeVP_Third = None
    InterAWSaleBeforeVP_Third = None
    InterUFSaleBeforeVP_Third = None
    InterUFLapseBeforeVP_Third = None
    LapseRiskOnExpBeforeVP_Third = None

    ProfitMarginPremAfterVP_Third = None
    CleanRiskPremiumAfterVP_Third = None
    initFixedPFeeAfterVP_Third = None
    renewalFixedPFeeAfterVP_Third = None
    claimsExpPremiumAfterVP_Third = None
    renewalVariPFeeAfterVP_Third = None
    DisribChannelSalesAfterVP_Third = None
    InterAWSaleAfterVP_Third = None
    InterUFSaleAfterVP_Third = None
    InterUFLapseAfterVP_Third = None
    LapseRiskOnExpAfterVP_Third = None

    ProfitMarginPremBeforeVP_Fourth = None
    CleanRiskPremiumBeforeVP_Fourth = None
    initFixedPFeeBeforeVP_Fourth = None
    renewalFixedPFeeBeforeVP_Fourth = None
    claimsExpPremiumBeforeVP_Fourth = None
    renewalVariPFeeBeforeVP_Fourth = None
    DisribChannelSalesBeforeVP_Fourth = None
    InterAWSaleBeforeVP_Fourth = None
    InterUFSaleBeforeVP_Fourth = None
    InterUFLapseBeforeVP_Fourth = None
    LapseRiskOnExpBeforeVP_Fourth = None

    ProfitMarginPremAfterVP_Fourth = None
    CleanRiskPremiumAfterVP_Fourth = None
    initFixedPFeeAfterVP_Fourth = None
    renewalFixedPFeeAfterVP_Fourth = None
    claimsExpPremiumAfterVP_Fourth = None
    renewalVariPFeeAfterVP_Fourth = None
    DisribChannelSalesAfterVP_Fourth = None
    InterAWSaleAfterVP_Fourth = None
    InterUFSaleAfterVP_Fourth = None
    InterUFLapseAfterVP_Fourth = None
    LapseRiskOnExpAfterVP_Fourth = None


# Creating object to read scenario lines
scenario_workbook = openpyxl.load_workbook(scenario_source)
scenario_cell = scenario_workbook.active
# Opening the pricing tool
tool_workbook = xlwings.Book(old_tool).sheets(3)
# Object to manipulate the pricing tool's calculations
xl_calc = xlwings.Book(old_tool)
# Object to write values to the scenario_source workbook
xl_write = xlwings.Book(scenario_source).sheets(1)


for x in range(8, 107):

    # Setting the workbook's calculation option to manual saves time
    xl_calc.app.calculation = 'manual'
    print("Scenario: ", x)

    # Reading the scenario and assigning input variables their values
    input.table_code_1 = scenario_cell.cell(row=x, column=15).value
    input.table_code_2 = scenario_cell.cell(row=x, column=53).value
    input.table_code_3 = scenario_cell.cell(row=x, column=91).value
    input.table_code_4 = scenario_cell.cell(row=x, column=129).value
    input.age = scenario_cell.cell(row=x, column=168).value
    input.dob = scenario_dob(Todays_date, input.age)
    input.sum_assured = scenario_cell.cell(row=x, column=32).value
    input.growth_Percentage = scenario_cell.cell(row=x, column=22).value
    input.package_code = str(scenario_cell.cell(row=x, column=27).value)

    # converting in order to copy to pricing tool
    if input.table_code_1 == "MFMM":
        input.table_code_1 = "Main life"
    elif input.table_code_1 == "MFSP":
        input.table_code_1 = "Spouse"
    elif input.table_code_1 == "MFOC":
        input.table_code_1 = "Child for Life"
    elif input.table_code_1 == "MFUC":
        input.table_code_1 = "Unlimited Children"
    elif input.table_code_1 == "MFPA":
        input.table_code_1 = "Parent"
    elif input.table_code_1 == "MFCL":
        input.table_code_1 = "Child for Life"
    elif input.table_code_1 == "MFEF":
        input.table_code_1 = "Extended Family"

    # Package codes and scheme codes
    if input.package_code == "801":
        input.scheme_code = "7005285"
    elif input.package_code == "802":
        input.scheme_code = "7005678"
    elif input.package_code == "803":
        input.scheme_code = "7005469"
    elif input.package_code == "804":
        input.scheme_code = "7005671"
    elif input.package_code == "805":
        input.scheme_code = "7005674"
    elif input.package_code == "806":
        input.scheme_code = "7005653"
    elif input.package_code == "807":
        input.scheme_code = "7005529"
    elif input.package_code == "808":
        input.scheme_code = "7005476"
    elif input.package_code == "809":
        input.scheme_code = "7005135"
    elif input.package_code == "810":
        input.scheme_code = "7005575"
    elif input.package_code == "811":
        input.scheme_code = "7005575"
    elif input.package_code == "812":
        input.scheme_code = "7005228"
    elif input.package_code == "813":
        input.scheme_code = "7005828"
    elif input.package_code == "814":
        input.scheme_code = "7005579"
    elif input.package_code == "815":
        input.scheme_code = "7005234"
    elif input.package_code == "816":
        input.scheme_code = "7005597"
    elif input.package_code == "817":
        input.scheme_code = "7005226"
    elif input.package_code == "818":
        input.scheme_code = "7001298"
    elif input.package_code == "819":
        input.scheme_code = "7005227"
    elif input.package_code == "820":
        input.scheme_code = "7005876"
    elif input.package_code == "841":
        input.scheme_code = "7005531"

    # Copying the values from the scenario sheet to the tool, used upper case 'Value' otherwise buffer error
    tool_workbook.cells(4, 3).value = input.table_code_1
    tool_workbook.cells(5, 3).value = input.table_code_2
    tool_workbook.cells(6, 3).value = input.table_code_3
    tool_workbook.cells(7, 3).value = input.table_code_4
    tool_workbook.cells(4, 4).value = input.dob
    tool_workbook.cells(4, 9).value = input.scheme_code
    tool_workbook.cells(4, 6).value = Todays_date
    tool_workbook.cells(4, 292).value = input.growth_Percentage/100

    # Giving all benefits the same sum assured even though the PUP's don't get sum assureds
    counter = 3
    for counter in range(4, 8):
        tool_workbook.cells(counter, 13).value = input.sum_assured

    # Selecting 'yes' or 'no' for the paid-ups
    if input.table_code_2 or input.table_code_3 or input.table_code_4 == "PUDS" or "PUDT":
        tool_workbook.cells(4, 14).value = "yes"
    if input.table_code_2 or input.table_code_3 or input.table_code_4 == "PURO" or "PURE":
        tool_workbook.cells(4, 15).value = "yes"

    xl_calc.app.calculation = 'automatic'

    # Giving output class variables their values, starting with the total premiums per benefit
    output.main_tot_before_VP = tool_workbook.cells(4, 17).value

    if input.table_code_2 == "PUDT":
        output.second_tot_before_VP = tool_workbook.cells(4, 29).value
    elif input.table_code_2 == "PUDS":
        output.second_tot_before_VP = tool_workbook.cells(4, 41).value
    elif input.table_code_2 == "PURO":
        output.second_tot_before_VP = tool_workbook.cells(4, 53).value
    elif input.table_code_2 == "PURE":
        output.second_tot_before_VP = tool_workbook.cells(4, 53).value
    else:
        output.second_tot_before_VP = 0

    if input.table_code_3 == "PUDT":
        output.third_tot_before_VP = tool_workbook.cells(4, 29).value
    elif input.table_code_3 == "PUDS":
        output.third_tot_before_VP = tool_workbook.cells(4, 41).value
    elif input.table_code_3 == "PURO":
        output.third_tot_before_VP = tool_workbook.cells(4, 53).value
    elif input.table_code_3 == "PURE":
        output.third_tot_before_VP = tool_workbook.cells(4, 53).value
    else:
        output.third_tot_before_VP = 0

    if input.table_code_4 == "PUDT":
        output.fourth_tot_before_VP = tool_workbook.cells(4, 29).value
    elif input.table_code_4 == "PUDS":
        output.fourth_tot_before_VP = tool_workbook.cells(4, 41).value
    elif input.table_code_4 == "PURO":
        output.fourth_tot_before_VP = tool_workbook.cells(4, 53).value
    elif input.table_code_4 == "PURE":
        output.fourth_tot_before_VP = tool_workbook.cells(4, 53).value
    else:
        output.fourth_tot_before_VP = 0

    # Elements for the Main benefit before VP
    output.ProfitMarginPremBeforeVP_Main = tool_workbook.cells(4, 18).value
    output.CleanRiskPremiumBeforeVP_Main = tool_workbook.cells(4, 19).value
    output.initFixedPFeeBeforeVP_Main = tool_workbook.cells(4, 20).value
    output.renewalFixedPFeeBeforeVP_Main = tool_workbook.cells(4, 21).value
    output.claimsExpPremiumBeforeVP_Main = tool_workbook.cells(4, 22).value
    output.renewalVariPFeeBeforeVP_Main = tool_workbook.cells(4, 23).value
    output.DisribChannelSalesBeforeVP_Main = tool_workbook.cells(4, 24).value
    output.InterAWSaleBeforeVP_Main = tool_workbook.cells(4, 25).value
    output.InterUFSaleBeforeVP_Main = tool_workbook.cells(4, 26).value
    output.InterUFLapseBeforeVP_Main = tool_workbook.cells(4, 27).value
    output.LapseRiskOnExpBeforeVP_Main = tool_workbook.cells(4, 28).value

    output.main_tot_elements_before_VP = mathtool.trunc(output.ProfitMarginPremBeforeVP_Main+output.CleanRiskPremiumBeforeVP_Main+output.initFixedPFeeBeforeVP_Main+
                                          output.renewalFixedPFeeBeforeVP_Main+output.claimsExpPremiumBeforeVP_Main+output.renewalVariPFeeBeforeVP_Main+
                                          output.DisribChannelSalesBeforeVP_Main+output.InterAWSaleBeforeVP_Main+output.InterUFSaleBeforeVP_Main+
                                          output.InterUFLapseBeforeVP_Main+output.LapseRiskOnExpBeforeVP_Main)

    # Elements for the second benefit before VP
    if input.table_code_2 == "PUDT":
        output.ProfitMarginPremBeforeVP_Second = tool_workbook.cells(4, 30).value
        output.CleanRiskPremiumBeforeVP_Second = tool_workbook.cells(4, 31).value
        output.initFixedPFeeBeforeVP_Second = tool_workbook.cells(4, 32).value
        output.renewalFixedPFeeBeforeVP_Second = tool_workbook.cells(4, 33).value
        output.claimsExpPremiumBeforeVP_Second = tool_workbook.cells(4, 34).value
        output.renewalVariPFeeBeforeVP_Second = tool_workbook.cells(4, 35).value
        output.DisribChannelSalesBeforeVP_Second = tool_workbook.cells(4, 36).value
        output.InterAWSaleBeforeVP_Second = tool_workbook.cells(4, 37).value
        output.InterUFSaleBeforeVP_Second = tool_workbook.cells(4, 38).value
        output.InterUFLapseBeforeVP_Second = tool_workbook.cells(4, 39).value
        output.LapseRiskOnExpBeforeVP_Second = tool_workbook.cells(4, 40).value

        output.pudt_tot_elements_before_VP = mathtool.trunc(output.ProfitMarginPremBeforeVP_Second+output.CleanRiskPremiumBeforeVP_Second+output.initFixedPFeeBeforeVP_Second+
                                                output.renewalFixedPFeeBeforeVP_Second+output.claimsExpPremiumBeforeVP_Second+output.renewalVariPFeeBeforeVP_Second+
                                                output.DisribChannelSalesBeforeVP_Second+output.InterAWSaleBeforeVP_Second+output.InterUFSaleBeforeVP_Second+
                                                output.InterUFLapseBeforeVP_Second+output.LapseRiskOnExpBeforeVP_Second)

    # Elements for the second benefit before VP
    elif input.table_code_2 == "PUDS":
        output.ProfitMarginPremBeforeVP_Second = tool_workbook.cells(4, 42).value
        output.CleanRiskPremiumBeforeVP_Second = tool_workbook.cells(4, 43).value
        output.initFixedPFeeBeforeVP_Second = tool_workbook.cells(4, 44).value
        output.renewalFixedPFeeBeforeVP_Second = tool_workbook.cells(4, 45).value
        output.claimsExpPremiumBeforeVP_Second = tool_workbook.cells(4, 46).value
        output.renewalVariPFeeBeforeVP_Second = tool_workbook.cells(4, 47).value
        output.DisribChannelSalesBeforeVP_Second = tool_workbook.cells(4, 48).value
        output.InterAWSaleBeforeVP_Second = tool_workbook.cells(4, 49).value
        output.InterUFSaleBeforeVP_Second = tool_workbook.cells(4, 50).value
        output.InterUFLapseBeforeVP_Second = tool_workbook.cells(4, 51).value
        output.LapseRiskOnExpBeforeVP_Second = tool_workbook.cells(4, 52).value

        output.puds_tot_elements_before_VP = mathtool.trunc(
                                            output.ProfitMarginPremBeforeVP_Second + output.CleanRiskPremiumBeforeVP_Second + output.initFixedPFeeBeforeVP_Second +
                                            output.renewalFixedPFeeBeforeVP_Second + output.claimsExpPremiumBeforeVP_Second + output.renewalVariPFeeBeforeVP_Second +
                                            output.DisribChannelSalesBeforeVP_Second + output.InterAWSaleBeforeVP_Second + output.InterUFSaleBeforeVP_Second +
                                            output.InterUFLapseBeforeVP_Second + output.LapseRiskOnExpBeforeVP_Second)

    # Elements for the second benefit before VP
    elif input.table_code_2 == "PURE":
        output.ProfitMarginPremBeforeVP_Second = tool_workbook.cells(4, 54).value
        output.CleanRiskPremiumBeforeVP_Second = tool_workbook.cells(4, 55).value
        output.initFixedPFeeBeforeVP_Second = tool_workbook.cells(4, 56).value
        output.renewalFixedPFeeBeforeVP_Second = tool_workbook.cells(4, 57).value
        output.claimsExpPremiumBeforeVP_Second = tool_workbook.cells(4, 58).value
        output.renewalVariPFeeBeforeVP_Second = tool_workbook.cells(4, 59).value
        output.DisribChannelSalesBeforeVP_Second = tool_workbook.cells(4, 60).value
        output.InterAWSaleBeforeVP_Second = tool_workbook.cells(4, 61).value
        output.InterUFSaleBeforeVP_Second = tool_workbook.cells(4, 62).value
        output.InterUFLapseBeforeVP_Second = tool_workbook.cells(4, 63).value
        output.LapseRiskOnExpBeforeVP_Second = tool_workbook.cells(4, 64).value

        output.pure_o_tot_elements_before_VP = mathtool.trunc(
                                                output.ProfitMarginPremBeforeVP_Second + output.CleanRiskPremiumBeforeVP_Second + output.initFixedPFeeBeforeVP_Second +
                                                output.renewalFixedPFeeBeforeVP_Second + output.claimsExpPremiumBeforeVP_Second + output.renewalVariPFeeBeforeVP_Second +
                                                output.DisribChannelSalesBeforeVP_Second + output.InterAWSaleBeforeVP_Second + output.InterUFSaleBeforeVP_Second +
                                                output.InterUFLapseBeforeVP_Second + output.LapseRiskOnExpBeforeVP_Second)

    elif input.table_code_2 == "PURO":
        output.ProfitMarginPremBeforeVP_Second = tool_workbook.cells(4, 54).value
        output.CleanRiskPremiumBeforeVP_Second = tool_workbook.cells(4, 55).value
        output.initFixedPFeeBeforeVP_Second = tool_workbook.cells(4, 56).value
        output.renewalFixedPFeeBeforeVP_Second = tool_workbook.cells(4, 57).value
        output.claimsExpPremiumBeforeVP_Second = tool_workbook.cells(4, 58).value
        output.renewalVariPFeeBeforeVP_Second = tool_workbook.cells(4, 59).value
        output.DisribChannelSalesBeforeVP_Second = tool_workbook.cells(4, 60).value
        output.InterAWSaleBeforeVP_Second = tool_workbook.cells(4, 61).value
        output.InterUFSaleBeforeVP_Second = tool_workbook.cells(4, 62).value
        output.InterUFLapseBeforeVP_Second = tool_workbook.cells(4, 63).value
        output.LapseRiskOnExpBeforeVP_Second = tool_workbook.cells(4, 64).value

        output.pure_o_tot_elements_before_VP = mathtool.trunc(
                                                output.ProfitMarginPremBeforeVP_Second + output.CleanRiskPremiumBeforeVP_Second + output.initFixedPFeeBeforeVP_Second +
                                                output.renewalFixedPFeeBeforeVP_Second + output.claimsExpPremiumBeforeVP_Second + output.renewalVariPFeeBeforeVP_Second +
                                                output.DisribChannelSalesBeforeVP_Second + output.InterAWSaleBeforeVP_Second + output.InterUFSaleBeforeVP_Second +
                                                output.InterUFLapseBeforeVP_Second + output.LapseRiskOnExpBeforeVP_Second)

    else:
        output.ProfitMarginPremBeforeVP_Second = 0
        output.CleanRiskPremiumBeforeVP_Second = 0
        output.initFixedPFeeBeforeVP_Second = 0
        output.renewalFixedPFeeBeforeVP_Second = 0
        output.claimsExpPremiumBeforeVP_Second = 0
        output.renewalVariPFeeBeforeVP_Second = 0
        output.DisribChannelSalesBeforeVP_Second = 0
        output.InterAWSaleBeforeVP_Second = 0
        output.InterUFSaleBeforeVP_Second = 0
        output.InterUFLapseBeforeVP_Second = 0
        output.LapseRiskOnExpBeforeVP_Second = 0

    # Getting elements for benefit 3
    if input.table_code_3 == "PUDT":
        output.ProfitMarginPremBeforeVP_Third = tool_workbook.cells(4, 30).value
        output.CleanRiskPremiumBeforeVP_Third = tool_workbook.cells(4, 31).value
        output.initFixedPFeeBeforeVP_Third = tool_workbook.cells(4, 32).value
        output.renewalFixedPFeeBeforeVP_Third = tool_workbook.cells(4, 33).value
        output.claimsExpPremiumBeforeVP_Third = tool_workbook.cells(4, 34).value
        output.renewalVariPFeeBeforeVP_Third = tool_workbook.cells(4, 35).value
        output.DisribChannelSalesBeforeVP_Third = tool_workbook.cells(4, 36).value
        output.InterAWSaleBeforeVP_Third = tool_workbook.cells(4, 37).value
        output.InterUFSaleBeforeVP_Third = tool_workbook.cells(4, 38).value
        output.InterUFLapseBeforeVP_Third = tool_workbook.cells(4, 39).value
        output.LapseRiskOnExpBeforeVP_Third = tool_workbook.cells(4, 40).value

        output.pudt_tot_elements_before_VP = mathtool.trunc(output.ProfitMarginPremBeforeVP_Third+output.CleanRiskPremiumBeforeVP_Third+output.initFixedPFeeBeforeVP_Third+
                                                output.renewalFixedPFeeBeforeVP_Third + output.claimsExpPremiumBeforeVP_Third + output.renewalVariPFeeBeforeVP_Third+
                                                output.DisribChannelSalesBeforeVP_Third + output.InterAWSaleBeforeVP_Third + output.InterUFSaleBeforeVP_Third +
                                                output.InterUFLapseBeforeVP_Third + output.LapseRiskOnExpBeforeVP_Third)

    elif input.table_code_3 == "PUDS":
        output.ProfitMarginPremBeforeVP_Third = tool_workbook.cells(4, 42).value
        output.CleanRiskPremiumBeforeVP_Third = tool_workbook.cells(4, 43).value
        output.initFixedPFeeBeforeVP_Third = tool_workbook.cells(4, 44).value
        output.renewalFixedPFeeBeforeVP_Third = tool_workbook.cells(4, 45).value
        output.claimsExpPremiumBeforeVP_Third = tool_workbook.cells(4, 46).value
        output.renewalVariPFeeBeforeVP_Third = tool_workbook.cells(4, 47).value
        output.DisribChannelSalesBeforeVP_Third = tool_workbook.cells(4, 48).value
        output.InterAWSaleBeforeVP_Third = tool_workbook.cells(4, 49).value
        output.InterUFSaleBeforeVP_Third = tool_workbook.cells(4, 50).value
        output.InterUFLapseBeforeVP_Third = tool_workbook.cells(4, 51).value
        output.LapseRiskOnExpBeforeVP_Third = tool_workbook.cells(4, 52).value

        output.puds_tot_elements_before_VP = mathtool.trunc(
                                                output.ProfitMarginPremBeforeVP_Third + output.CleanRiskPremiumBeforeVP_Third + output.initFixedPFeeBeforeVP_Third+
                                                output.renewalFixedPFeeBeforeVP_Third + output.claimsExpPremiumBeforeVP_Third + output.renewalVariPFeeBeforeVP_Third +
                                                output.DisribChannelSalesBeforeVP_Third + output.InterAWSaleBeforeVP_Third + output.InterUFSaleBeforeVP_Third +
                                                output.InterUFLapseBeforeVP_Third + output.LapseRiskOnExpBeforeVP_Third)

    elif input.table_code_3 == "PURE":
        output.ProfitMarginPremBeforeVP_Third = tool_workbook.cells(4, 54).value
        output.CleanRiskPremiumBeforeVP_Third = tool_workbook.cells(4, 55).value
        output.initFixedPFeeBeforeVP_Third = tool_workbook.cells(4, 56).value
        output.renewalFixedPFeeBeforeVP_Third = tool_workbook.cells(4, 57).value
        output.claimsExpPremiumBeforeVP_Third = tool_workbook.cells(4, 58).value
        output.renewalVariPFeeBeforeVP_Third = tool_workbook.cells(4, 59).value
        output.DisribChannelSalesBeforeVP_Third = tool_workbook.cells(4, 60).value
        output.InterAWSaleBeforeVP_Third = tool_workbook.cells(4, 61).value
        output.InterUFSaleBeforeVP_Third = tool_workbook.cells(4, 62).value
        output.InterUFLapseBeforeVP_Third = tool_workbook.cells(4, 63).value
        output.LapseRiskOnExpBeforeVP_Third = tool_workbook.cells(4, 64).value

        output.pure_o_tot_elements_before_VP = mathtool.trunc(
                                                output.ProfitMarginPremBeforeVP_Third + output.CleanRiskPremiumBeforeVP_Third + output.initFixedPFeeBeforeVP_Third +
                                                output.renewalFixedPFeeBeforeVP_Third + output.claimsExpPremiumBeforeVP_Third + output.renewalVariPFeeBeforeVP_Third +
                                                output.DisribChannelSalesBeforeVP_Third + output.InterAWSaleBeforeVP_Third + output.InterUFSaleBeforeVP_Third +
                                                output.InterUFLapseBeforeVP_Third + output.LapseRiskOnExpBeforeVP_Third)

    elif input.table_code_3 == "PURO":
        output.ProfitMarginPremBeforeVP_Third = tool_workbook.cells(4, 54).value
        output.CleanRiskPremiumBeforeVP_Third = tool_workbook.cells(4, 55).value
        output.initFixedPFeeBeforeVP_Third = tool_workbook.cells(4, 56).value
        output.renewalFixedPFeeBeforeVP_Third = tool_workbook.cells(4, 57).value
        output.claimsExpPremiumBeforeVP_Third = tool_workbook.cells(4, 58).value
        output.renewalVariPFeeBeforeVP_Third = tool_workbook.cells(4, 59).value
        output.DisribChannelSalesBeforeVP_Third = tool_workbook.cells(4, 60).value
        output.InterAWSaleBeforeVP_Third = tool_workbook.cells(4, 61).value
        output.InterUFSaleBeforeVP_Third = tool_workbook.cells(4, 62).value
        output.InterUFLapseBeforeVP_Third = tool_workbook.cells(4, 63).value
        output.LapseRiskOnExpBeforeVP_Third = tool_workbook.cells(4, 64).value

        output.pure_o_tot_elements_before_VP = mathtool.trunc(
                                                output.ProfitMarginPremBeforeVP_Third + output.CleanRiskPremiumBeforeVP_Third + output.initFixedPFeeBeforeVP_Third +
                                                output.renewalFixedPFeeBeforeVP_Third + output.claimsExpPremiumBeforeVP_Third + output.renewalVariPFeeBeforeVP_Third +
                                                output.DisribChannelSalesBeforeVP_Third + output.InterAWSaleBeforeVP_Third + output.InterUFSaleBeforeVP_Third +
                                                output.InterUFLapseBeforeVP_Third + output.LapseRiskOnExpBeforeVP_Third)

    else:
        output.ProfitMarginPremBeforeVP_Third = 0
        output.CleanRiskPremiumBeforeVP_Third = 0
        output.initFixedPFeeBeforeVP_Third = 0
        output.renewalFixedPFeeBeforeVP_Third = 0
        output.claimsExpPremiumBeforeVP_Third = 0
        output.renewalVariPFeeBeforeVP_Third = 0
        output.DisribChannelSalesBeforeVP_Third= 0
        output.InterAWSaleBeforeVP_Third = 0
        output.InterUFSaleBeforeVP_Third = 0
        output.InterUFLapseBeforeVP_Third = 0
        output.LapseRiskOnExpBeforeVP_Third = 0

    # Getting elements for benefit 4
    if input.table_code_4 == "PUDT":
        output.ProfitMarginPremBeforeVP_Fourth = tool_workbook.cells(4, 30).value
        output.CleanRiskPremiumBeforeVP_Fourth = tool_workbook.cells(4, 31).value
        output.initFixedPFeeBeforeVP_Fourth = tool_workbook.cells(4, 32).value
        output.renewalFixedPFeeBeforeVP_Fourth = tool_workbook.cells(4, 33).value
        output.claimsExpPremiumBeforeVP_Fourth = tool_workbook.cells(4, 34).value
        output.renewalVariPFeeBeforeVP_Fourth = tool_workbook.cells(4, 35).value
        output.DisribChannelSalesBeforeVP_Fourth = tool_workbook.cells(4, 36).value
        output.InterAWSaleBeforeVP_Fourth = tool_workbook.cells(4, 37).value
        output.InterUFSaleBeforeVP_Fourth = tool_workbook.cells(4, 38).value
        output.InterUFLapseBeforeVP_Fourth = tool_workbook.cells(4, 39).value
        output.LapseRiskOnExpBeforeVP_Fourth = tool_workbook.cells(4, 40).value

        output.pudt_tot_elements_before_VP = mathtool.trunc(
                                            output.ProfitMarginPremBeforeVP_Fourth + output.CleanRiskPremiumBeforeVP_Fourth + output.initFixedPFeeBeforeVP_Fourth+
                                            output.renewalFixedPFeeBeforeVP_Fourth + output.claimsExpPremiumBeforeVP_Fourth + output.renewalVariPFeeBeforeVP_Fourth+
                                            output.DisribChannelSalesBeforeVP_Fourth + output.InterAWSaleBeforeVP_Fourth + output.InterUFSaleBeforeVP_Fourth +
                                            output.InterUFLapseBeforeVP_Fourth + output.LapseRiskOnExpBeforeVP_Fourth)

    elif input.table_code_4 == "PUDS":
        output.ProfitMarginPremBeforeVP_Fourth = tool_workbook.cells(4, 42).value
        output.CleanRiskPremiumBeforeVP_Fourth = tool_workbook.cells(4, 43).value
        output.initFixedPFeeBeforeVP_Fourth = tool_workbook.cells(4, 44).value
        output.renewalFixedPFeeBeforeVP_Fourth = tool_workbook.cells(4, 45).value
        output.claimsExpPremiumBeforeVP_Fourth = tool_workbook.cells(4, 46).value
        output.renewalVariPFeeBeforeVP_Fourth = tool_workbook.cells(4, 47).value
        output.DisribChannelSalesBeforeVP_Fourth = tool_workbook.cells(4, 48).value
        output.InterAWSaleBeforeVP_Fourth = tool_workbook.cells(4, 49).value
        output.InterUFSaleBeforeVP_Fourth = tool_workbook.cells(4, 50).value
        output.InterUFLapseBeforeVP_Fourth = tool_workbook.cells(4, 51).value
        output.LapseRiskOnExpBeforeVP_Fourth = tool_workbook.cells(4, 52).value

        output.puds_tot_elements_before_VP = mathtool.trunc(
                                                output.ProfitMarginPremBeforeVP_Fourth + output.CleanRiskPremiumBeforeVP_Fourth + output.initFixedPFeeBeforeVP_Fourth+
                                                output.renewalFixedPFeeBeforeVP_Fourth + output.claimsExpPremiumBeforeVP_Fourth + output.renewalVariPFeeBeforeVP_Fourth +
                                                output.DisribChannelSalesBeforeVP_Fourth + output.InterAWSaleBeforeVP_Fourth + output.InterUFSaleBeforeVP_Fourth +
                                                output.InterUFLapseBeforeVP_Fourth + output.LapseRiskOnExpBeforeVP_Fourth)

    elif input.table_code_4 == "PURE":
        output.ProfitMarginPremBeforeVP_Fourth = tool_workbook.cells(4, 54).value
        output.CleanRiskPremiumBeforeVP_Fourth = tool_workbook.cells(4, 55).value
        output.initFixedPFeeBeforeVP_Fourth = tool_workbook.cells(4, 56).value
        output.renewalFixedPFeeBeforeVP_Fourth = tool_workbook.cells(4, 57).value
        output.claimsExpPremiumBeforeVP_Fourth = tool_workbook.cells(4, 58).value
        output.renewalVariPFeeBeforeVP_Fourth = tool_workbook.cells(4, 59).value
        output.DisribChannelSalesBeforeVP_Fourth = tool_workbook.cells(4, 60).value
        output.InterAWSaleBeforeVP_Fourth = tool_workbook.cells(4, 61).value
        output.InterUFSaleBeforeVP_Fourth = tool_workbook.cells(4, 62).value
        output.InterUFLapseBeforeVP_Fourth = tool_workbook.cells(4, 63).value
        output.LapseRiskOnExpBeforeVP_Fourth = tool_workbook.cells(4, 64).value

        output.pure_o_tot_elements_before_VP = mathtool.trunc(
                                                output.ProfitMarginPremBeforeVP_Fourth + output.CleanRiskPremiumBeforeVP_Fourth + output.initFixedPFeeBeforeVP_Fourth+
                                                output.renewalFixedPFeeBeforeVP_Fourth + output.claimsExpPremiumBeforeVP_Fourth + output.renewalVariPFeeBeforeVP_Fourth +
                                                output.DisribChannelSalesBeforeVP_Fourth + output.InterAWSaleBeforeVP_Fourth + output.InterUFSaleBeforeVP_Fourth +
                                                output.InterUFLapseBeforeVP_Fourth + output.LapseRiskOnExpBeforeVP_Fourth)
    elif input.table_code_4 == "PURO":
        output.ProfitMarginPremBeforeVP_Fourth = tool_workbook.cells(4, 54).value
        output.CleanRiskPremiumBeforeVP_Fourth = tool_workbook.cells(4, 55).value
        output.initFixedPFeeBeforeVP_Fourth = tool_workbook.cells(4, 56).value
        output.renewalFixedPFeeBeforeVP_Fourth = tool_workbook.cells(4, 57).value
        output.claimsExpPremiumBeforeVP_Fourth = tool_workbook.cells(4, 58).value
        output.renewalVariPFeeBeforeVP_Fourth = tool_workbook.cells(4, 59).value
        output.DisribChannelSalesBeforeVP_Fourth = tool_workbook.cells(4, 60).value
        output.InterAWSaleBeforeVP_Fourth = tool_workbook.cells(4, 61).value
        output.InterUFSaleBeforeVP_Fourth = tool_workbook.cells(4, 62).value
        output.InterUFLapseBeforeVP_Fourth = tool_workbook.cells(4, 63).value
        output.LapseRiskOnExpBeforeVP_Fourth = tool_workbook.cells(4, 64).value

        output.pure_o_tot_elements_before_VP = mathtool.trunc(
                                                output.ProfitMarginPremBeforeVP_Fourth + output.CleanRiskPremiumBeforeVP_Fourth + output.initFixedPFeeBeforeVP_Fourth +
                                                output.renewalFixedPFeeBeforeVP_Fourth + output.claimsExpPremiumBeforeVP_Fourth + output.renewalVariPFeeBeforeVP_Fourth +
                                                output.DisribChannelSalesBeforeVP_Fourth + output.InterAWSaleBeforeVP_Fourth + output.InterUFSaleBeforeVP_Fourth +
                                                output.InterUFLapseBeforeVP_Fourth + output.LapseRiskOnExpBeforeVP_Fourth)

    else:
        output.ProfitMarginPremBeforeVP_Fourth = 0
        output.CleanRiskPremiumBeforeVP_Fourth = 0
        output.initFixedPFeeBeforeVP_Fourth = 0
        output.renewalFixedPFeeBeforeVP_Fourth = 0
        output.claimsExpPremiumBeforeVP_Fourth = 0
        output.renewalVariPFeeBeforeVP_Fourth = 0
        output.DisribChannelSalesBeforeVP_Fourth = 0
        output.InterAWSaleBeforeVP_Fourth = 0
        output.InterUFSaleBeforeVP_Fourth = 0
        output.InterUFLapseBeforeVP_Fourth = 0
        output.LapseRiskOnExpBeforeVP_Fourth = 0

    output.main_tot_after_VP = tool_workbook.cells(4, 346).value

    if input.table_code_2 == "PUDT":
        output.second_tot_after_VP = tool_workbook.cells(4, 358).value
    elif input.table_code_2 == "PUDS":
        output.second_tot_after_VP = tool_workbook.cells(4, 370).value
    elif input.table_code_2 == "PURE":
        output.second_tot_after_VP = tool_workbook.cells(4, 382).value
    elif input.table_code_2 == "PURO":
        output.second_tot_after_VP = tool_workbook.cells(4, 382).value
    else:
        output.second_tot_after_VP = 0

    if input.table_code_3 == "PUDT":
        output.third_tot_after_VP = tool_workbook.cells(4, 358).value
    elif input.table_code_3 == "PUDS":
        output.third_tot_after_VP = tool_workbook.cells(4, 370).value
    elif input.table_code_3 == "PURE":
        output.third_tot_after_VP = tool_workbook.cells(4, 382).value
    elif input.table_code_3 == "PURO":
        output.third_tot_after_VP = tool_workbook.cells(4, 382).value
    else:
        output.third_tot_after_VP = 0

    if input.table_code_4 == "PUDT":
        output.fourth_tot_after_VP = tool_workbook.cells(4, 358).value
    elif input.table_code_4 == "PUDS":
        output.fourth_tot_after_VP = tool_workbook.cells(4, 370).value
    elif input.table_code_4 == "PURE":
        output.fourth_tot_after_VP = tool_workbook.cells(4, 382).value
    elif input.table_code_4 == "PURO":
        output.fourth_tot_after_VP = tool_workbook.cells(4, 382).value
    else:
        output.fourth_tot_after_VP = 0

    # Elements for the Main benefit after VP
    output.ProfitMarginPremAfterVP_Main = tool_workbook.cells(4, 347).value
    output.CleanRiskPremiumAfterVP_Main = tool_workbook.cells(4, 348).value
    output.initFixedPFeeAfterVP_Main = tool_workbook.cells(4, 349).value
    output.renewalFixedPFeeAfterVP_Main = tool_workbook.cells(4, 350).value
    output.claimsExpPremiumAfterVP_Main = tool_workbook.cells(4, 351).value
    output.renewalVariPFeeAfterVP_Main = tool_workbook.cells(4, 352).value
    output.DisribChannelSalesAfterVP_Main = tool_workbook.cells(4, 353).value
    output.InterAWSaleAfterVP_Main = tool_workbook.cells(4, 354).value
    output.InterUFSaleAfterVP_Main = tool_workbook.cells(4, 355).value
    output.InterUFLapseAfterVP_Main = tool_workbook.cells(4, 356).value
    output.LapseRiskOnExpAfterVP_Main = tool_workbook.cells(4, 357).value

    output.main_tot_elements_after_VP = mathtool.trunc(output.ProfitMarginPremAfterVP_Main+output.CleanRiskPremiumAfterVP_Main+output.initFixedPFeeAfterVP_Main+
                                                       output.renewalFixedPFeeAfterVP_Main+output.claimsExpPremiumAfterVP_Main+output.renewalVariPFeeAfterVP_Main+
                                                       output.DisribChannelSalesAfterVP_Main+output.InterAWSaleAfterVP_Main+output.InterUFSaleAfterVP_Main+
                                                       output.InterUFLapseAfterVP_Main+output.LapseRiskOnExpAfterVP_Main)

    # Elements for the 'PUDT' benefit after VP
    if input.table_code_2 == "PUDT":
        output.ProfitMarginPremAfterVP_Second = tool_workbook.cells(4, 359).value
        output.CleanRiskPremiumAfterVP_Second = tool_workbook.cells(4, 360).value
        output.initFixedPFeeAfterVP_Second = tool_workbook.cells(4, 361).value
        output.renewalFixedPFeeAfterVP_Second = tool_workbook.cells(4, 362).value
        output.claimsExpPremiumAfterVP_Second = tool_workbook.cells(4, 363).value
        output.renewalVariPFeeAfterVP_Second = tool_workbook.cells(4, 364).value
        output.DisribChannelSalesAfterVP_Second = tool_workbook.cells(4, 365).value
        output.InterAWSaleAfterVP_Second = tool_workbook.cells(4, 366).value
        output.InterUFSaleAfterVP_Second = tool_workbook.cells(4, 367).value
        output.InterUFLapseAfterVP_Second = tool_workbook.cells(4, 368).value
        output.LapseRiskOnExpAfterVP_Second = tool_workbook.cells(4, 369).value

        output.pudt_tot_elements_after_VP = mathtool.trunc(output.ProfitMarginPremAfterVP_Second+output.CleanRiskPremiumAfterVP_Second+output.initFixedPFeeAfterVP_Second+
                                                       output.renewalFixedPFeeAfterVP_Second+output.claimsExpPremiumAfterVP_Second+output.renewalVariPFeeAfterVP_Second+
                                                       output.DisribChannelSalesAfterVP_Second+output.InterAWSaleAfterVP_Second+output.InterUFSaleAfterVP_Second+
                                                       output.InterUFLapseAfterVP_Second+output.LapseRiskOnExpAfterVP_Second)

    # Elements for the 'PUDS' benefit after VP
    elif input.table_code_2 == "PUDS":
        output.ProfitMarginPremAfterVP_Second = tool_workbook.cells(4, 371).value
        output.CleanRiskPremiumAfterVP_Second = tool_workbook.cells(4, 372).value
        output.initFixedPFeeAfterVP_Second = tool_workbook.cells(4, 373).value
        output.renewalFixedPFeeAfterVP_Second = tool_workbook.cells(4, 374).value
        output.claimsExpPremiumAfterVP_Second = tool_workbook.cells(4, 375).value
        output.renewalVariPFeeAfterVP_Second = tool_workbook.cells(4, 376).value
        output.DisribChannelSalesAfterVP_Second = tool_workbook.cells(4, 377).value
        output.InterAWSaleAfterVP_Second = tool_workbook.cells(4, 378).value
        output.InterUFSaleAfterVP_Second = tool_workbook.cells(4, 379).value
        output.InterUFLapseAfterVP_Second = tool_workbook.cells(4, 380).value
        output.LapseRiskOnExpAfterVP_Second = tool_workbook.cells(4, 381).value

        output.puds_tot_elements_after_VP = mathtool.trunc(
                                            output.ProfitMarginPremAfterVP_Second + output.CleanRiskPremiumAfterVP_Second + output.initFixedPFeeAfterVP_Second +
                                            output.renewalFixedPFeeAfterVP_Second + output.claimsExpPremiumAfterVP_Second + output.renewalVariPFeeAfterVP_Second +
                                            output.DisribChannelSalesAfterVP_Second + output.InterAWSaleAfterVP_Second + output.InterUFSaleAfterVP_Second +
                                            output.InterUFLapseAfterVP_Second + output.LapseRiskOnExpAfterVP_Second)

    # Elements for the 'PURE/PURO' benefit after VP
    elif input.table_code_2 == "PURE":
        output.ProfitMarginPremAfterVP_Second = tool_workbook.cells(4, 383).value
        output.CleanRiskPremiumAfterVP_Second = tool_workbook.cells(4, 384).value
        output.initFixedPFeeAfterVP_Second = tool_workbook.cells(4, 385).value
        output.renewalFixedPFeeAfterVP_Second = tool_workbook.cells(4, 386).value
        output.claimsExpPremiumAfterVP_Second = tool_workbook.cells(4, 387).value
        output.renewalVariPFeeAfterVP_Second = tool_workbook.cells(4, 388).value
        output.DisribChannelSalesAfterVP_Second = tool_workbook.cells(4, 389).value
        output.InterAWSaleAfterVP_Second = tool_workbook.cells(4, 390).value
        output.InterUFSaleAfterVP_Second = tool_workbook.cells(4, 391).value
        output.InterUFLapseAfterVP_Second = tool_workbook.cells(4, 392).value
        output.LapseRiskOnExpAfterVP_Second = tool_workbook.cells(4, 393).value

        output.pure_o_tot_elements_after_VP = mathtool.trunc(
                                            output.ProfitMarginPremAfterVP_Second + output.CleanRiskPremiumAfterVP_Second + output.initFixedPFeeAfterVP_Second +
                                            output.renewalFixedPFeeAfterVP_Second + output.claimsExpPremiumAfterVP_Second + output.renewalVariPFeeAfterVP_Second +
                                            output.DisribChannelSalesAfterVP_Second + output.InterAWSaleAfterVP_Second + output.InterUFSaleAfterVP_Second +
                                            output.InterUFLapseAfterVP_Second + output.LapseRiskOnExpAfterVP_Second)

    elif input.table_code_2 == "PURO":
        output.ProfitMarginPremAfterVP_Second = tool_workbook.cells(4, 383).value
        output.CleanRiskPremiumAfterVP_Second = tool_workbook.cells(4, 384).value
        output.initFixedPFeeAfterVP_Second = tool_workbook.cells(4, 385).value
        output.renewalFixedPFeeAfterVP_Second = tool_workbook.cells(4, 386).value
        output.claimsExpPremiumAfterVP_Second = tool_workbook.cells(4, 387).value
        output.renewalVariPFeeAfterVP_Second = tool_workbook.cells(4, 388).value
        output.DisribChannelSalesAfterVP_Second = tool_workbook.cells(4, 389).value
        output.InterAWSaleAfterVP_Second = tool_workbook.cells(4, 390).value
        output.InterUFSaleAfterVP_Second = tool_workbook.cells(4, 391).value
        output.InterUFLapseAfterVP_Second = tool_workbook.cells(4, 392).value
        output.LapseRiskOnExpAfterVP_Second = tool_workbook.cells(4, 393).value

        output.pure_o_tot_elements_after_VP = mathtool.trunc(
                                            output.ProfitMarginPremAfterVP_Second + output.CleanRiskPremiumAfterVP_Second + output.initFixedPFeeAfterVP_Second +
                                            output.renewalFixedPFeeAfterVP_Second + output.claimsExpPremiumAfterVP_Second + output.renewalVariPFeeAfterVP_Second +
                                            output.DisribChannelSalesAfterVP_Second + output.InterAWSaleAfterVP_Second + output.InterUFSaleAfterVP_Second +
                                            output.InterUFLapseAfterVP_Second + output.LapseRiskOnExpAfterVP_Second)

    else:
        output.ProfitMarginPremAfterVP_Second = 0
        output.CleanRiskPremiumAfterVP_Second = 0
        output.initFixedPFeeAfterVP_Second = 0
        output.renewalFixedPFeeAfterVP_Second = 0
        output.claimsExpPremiumAfterVP_Second = 0
        output.renewalVariPFeeAfterVP_Second = 0
        output.DisribChannelSalesAfterVP_Second = 0
        output.InterAWSaleAfterVP_Second = 0
        output.InterUFSaleAfterVP_Second = 0
        output.InterUFLapseAfterVP_Second = 0
        output.LapseRiskOnExpAfterVP_Second = 0

    # Getting elements after VP for the third benefit
    if input.table_code_3 == "PUDT":
        output.ProfitMarginPremAfterVP_Third = tool_workbook.cells(4, 359).value
        output.CleanRiskPremiumAfterVP_Third = tool_workbook.cells(4, 360).value
        output.initFixedPFeeAfterVP_Third = tool_workbook.cells(4, 361).value
        output.renewalFixedPFeeAfterVP_Third = tool_workbook.cells(4, 362).value
        output.claimsExpPremiumAfterVP_Third = tool_workbook.cells(4, 363).value
        output.renewalVariPFeeAfterVP_Third = tool_workbook.cells(4, 364).value
        output.DisribChannelSalesAfterVP_Third = tool_workbook.cells(4, 365).value
        output.InterAWSaleAfterVP_Third = tool_workbook.cells(4, 366).value
        output.InterUFSaleAfterVP_Third = tool_workbook.cells(4, 367).value
        output.InterUFLapseAfterVP_Third = tool_workbook.cells(4, 368).value
        output.LapseRiskOnExpAfterVP_Third = tool_workbook.cells(4, 369).value

        output.pudt_tot_elements_after_VP = mathtool.trunc(
                                                        output.ProfitMarginPremAfterVP_Third + output.CleanRiskPremiumAfterVP_Third + output.initFixedPFeeAfterVP_Third +
                                                        output.renewalFixedPFeeAfterVP_Third + output.claimsExpPremiumAfterVP_Third + output.renewalVariPFeeAfterVP_Third +
                                                        output.DisribChannelSalesAfterVP_Third + output.InterAWSaleAfterVP_Third + output.InterUFSaleAfterVP_Third +
                                                        output.InterUFLapseAfterVP_Third + output.LapseRiskOnExpAfterVP_Third)

    elif input.table_code_3 == "PUDS":
        output.ProfitMarginPremAfterVP_Third = tool_workbook.cells(4, 371).value
        output.CleanRiskPremiumAfterVP_Third = tool_workbook.cells(4, 372).value
        output.initFixedPFeeAfterVP_Third = tool_workbook.cells(4, 373).value
        output.renewalFixedPFeeAfterVP_Third = tool_workbook.cells(4, 374).value
        output.claimsExpPremiumAfterVP_Third = tool_workbook.cells(4, 375).value
        output.renewalVariPFeeAfterVP_Third = tool_workbook.cells(4, 376).value
        output.DisribChannelSalesAfterVP_Third = tool_workbook.cells(4, 377).value
        output.InterAWSaleAfterVP_Third = tool_workbook.cells(4, 378).value
        output.InterUFSaleAfterVP_Third = tool_workbook.cells(4, 379).value
        output.InterUFLapseAfterVP_Third = tool_workbook.cells(4, 380).value
        output.LapseRiskOnExpAfterVP_Third = tool_workbook.cells(4, 381).value

        output.puds_tot_elements_after_VP = mathtool.trunc(
                                                        output.ProfitMarginPremAfterVP_Third + output.CleanRiskPremiumAfterVP_Third + output.initFixedPFeeAfterVP_Third +
                                                        output.renewalFixedPFeeAfterVP_Third + output.claimsExpPremiumAfterVP_Third + output.renewalVariPFeeAfterVP_Third +
                                                        output.DisribChannelSalesAfterVP_Third + output.InterAWSaleAfterVP_Third + output.InterUFSaleAfterVP_Third +
                                                        output.InterUFLapseAfterVP_Third + output.LapseRiskOnExpAfterVP_Third)

    elif input.table_code_3 == "PURE":
        output.ProfitMarginPremAfterVP_Third = tool_workbook.cells(4, 383).value
        output.CleanRiskPremiumAfterVP_Third = tool_workbook.cells(4, 384).value
        output.initFixedPFeeAfterVP_Third = tool_workbook.cells(4, 385).value
        output.renewalFixedPFeeAfterVP_Third = tool_workbook.cells(4, 386).value
        output.claimsExpPremiumAfterVP_Third = tool_workbook.cells(4, 387).value
        output.renewalVariPFeeAfterVP_Third = tool_workbook.cells(4, 388).value
        output.DisribChannelSalesAfterVP_Third = tool_workbook.cells(4, 389).value
        output.InterAWSaleAfterVP_Third = tool_workbook.cells(4, 390).value
        output.InterUFSaleAfterVP_Third = tool_workbook.cells(4, 391).value
        output.InterUFLapseAfterVP_Third = tool_workbook.cells(4, 392).value
        output.LapseRiskOnExpAfterVP_Third = tool_workbook.cells(4, 393).value

        output.pure_o_tot_elements_after_VP = mathtool.trunc(
                                                        output.ProfitMarginPremAfterVP_Third + output.CleanRiskPremiumAfterVP_Third + output.initFixedPFeeAfterVP_Third +
                                                        output.renewalFixedPFeeAfterVP_Third + output.claimsExpPremiumAfterVP_Third + output.renewalVariPFeeAfterVP_Third +
                                                        output.DisribChannelSalesAfterVP_Third + output.InterAWSaleAfterVP_Third + output.InterUFSaleAfterVP_Third +
                                                        output.InterUFLapseAfterVP_Third + output.LapseRiskOnExpAfterVP_Third)

    elif input.table_code_3 == "PURO":
        output.ProfitMarginPremAfterVP_Third = tool_workbook.cells(4, 383).value
        output.CleanRiskPremiumAfterVP_Third = tool_workbook.cells(4, 384).value
        output.initFixedPFeeAfterVP_Third = tool_workbook.cells(4, 385).value
        output.renewalFixedPFeeAfterVP_Third = tool_workbook.cells(4, 386).value
        output.claimsExpPremiumAfterVP_Third = tool_workbook.cells(4, 387).value
        output.renewalVariPFeeAfterVP_Third = tool_workbook.cells(4, 388).value
        output.DisribChannelSalesAfterVP_Third = tool_workbook.cells(4, 389).value
        output.InterAWSaleAfterVP_Third = tool_workbook.cells(4, 390).value
        output.InterUFSaleAfterVP_Third = tool_workbook.cells(4, 391).value
        output.InterUFLapseAfterVP_Third = tool_workbook.cells(4, 392).value
        output.LapseRiskOnExpAfterVP_Third = tool_workbook.cells(4, 393).value

        output.pure_o_tot_elements_after_VP = mathtool.trunc(
                                                output.ProfitMarginPremAfterVP_Third + output.CleanRiskPremiumAfterVP_Third + output.initFixedPFeeAfterVP_Third +
                                                output.renewalFixedPFeeAfterVP_Third + output.claimsExpPremiumAfterVP_Third + output.renewalVariPFeeAfterVP_Third +
                                                output.DisribChannelSalesAfterVP_Third + output.InterAWSaleAfterVP_Third + output.InterUFSaleAfterVP_Third +
                                                output.InterUFLapseAfterVP_Third + output.LapseRiskOnExpAfterVP_Third)

    else:
        output.ProfitMarginPremAfterVP_Third = 0
        output.CleanRiskPremiumAfterVP_Third = 0
        output.initFixedPFeeAfterVP_Third = 0
        output.renewalFixedPFeeAfterVP_Third = 0
        output.claimsExpPremiumAfterVP_Third = 0
        output.renewalVariPFeeAfterVP_Third = 0
        output.DisribChannelSalesAfterVP_Third = 0
        output.InterAWSaleAfterVP_Third = 0
        output.InterUFSaleAfterVP_Third = 0
        output.InterUFLapseAfterVP_Third = 0
        output.LapseRiskOnExpAfterVP_Third = 0

    # Getting elements for fourth benefit
    if input.table_code_4 == "PUDT":
        output.ProfitMarginPremAfterVP_Fourth = tool_workbook.cells(4, 359).value
        output.CleanRiskPremiumAfterVP_Fourth = tool_workbook.cells(4, 360).value
        output.initFixedPFeeAfterVP_Fourth = tool_workbook.cells(4, 361).value
        output.renewalFixedPFeeAfterVP_Fourth = tool_workbook.cells(4, 362).value
        output.claimsExpPremiumAfterVP_Fourth = tool_workbook.cells(4, 363).value
        output.renewalVariPFeeAfterVP_Fourth = tool_workbook.cells(4, 364).value
        output.DisribChannelSalesAfterVP_Fourth = tool_workbook.cells(4, 365).value
        output.InterAWSaleAfterVP_Fourth = tool_workbook.cells(4, 366).value
        output.InterUFSaleAfterVP_Fourth = tool_workbook.cells(4, 367).value
        output.InterUFLapseAfterVP_Fourth = tool_workbook.cells(4, 368).value
        output.LapseRiskOnExpAfterVP_Fourth = tool_workbook.cells(4, 369).value

        output.pudt_tot_elements_after_VP = mathtool.trunc(
                                                        output.ProfitMarginPremAfterVP_Fourth + output.CleanRiskPremiumAfterVP_Fourth + output.initFixedPFeeAfterVP_Fourth +
                                                        output.renewalFixedPFeeAfterVP_Fourth + output.claimsExpPremiumAfterVP_Fourth + output.renewalVariPFeeAfterVP_Fourth +
                                                        output.DisribChannelSalesAfterVP_Fourth + output.InterAWSaleAfterVP_Fourth + output.InterUFSaleAfterVP_Fourth +
                                                        output.InterUFLapseAfterVP_Fourth + output.LapseRiskOnExpAfterVP_Fourth)

    elif input.table_code_4 == "PUDS":
        output.ProfitMarginPremAfterVP_Fourth = tool_workbook.cells(4, 371).value
        output.CleanRiskPremiumAfterVP_Fourth = tool_workbook.cells(4, 372).value
        output.initFixedPFeeAfterVP_Fourth = tool_workbook.cells(4, 373).value
        output.renewalFixedPFeeAfterVP_Fourth = tool_workbook.cells(4, 374).value
        output.claimsExpPremiumAfterVP_Fourth = tool_workbook.cells(4, 375).value
        output.renewalVariPFeeAfterVP_Fourth = tool_workbook.cells(4, 376).value
        output.DisribChannelSalesAfterVP_Fourth = tool_workbook.cells(4, 377).value
        output.InterAWSaleAfterVP_Fourth = tool_workbook.cells(4, 378).value
        output.InterUFSaleAfterVP_Fourth = tool_workbook.cells(4, 379).value
        output.InterUFLapseAfterVP_Fourth = tool_workbook.cells(4, 380).value
        output.LapseRiskOnExpAfterVP_Fourth = tool_workbook.cells(4, 381).value

        output.puds_tot_elements_after_VP = mathtool.trunc(
                                                        output.ProfitMarginPremAfterVP_Fourth + output.CleanRiskPremiumAfterVP_Fourth + output.initFixedPFeeAfterVP_Fourth +
                                                        output.renewalFixedPFeeAfterVP_Fourth + output.claimsExpPremiumAfterVP_Fourth + output.renewalVariPFeeAfterVP_Fourth +
                                                        output.DisribChannelSalesAfterVP_Fourth + output.InterAWSaleAfterVP_Fourth + output.InterUFSaleAfterVP_Fourth +
                                                        output.InterUFLapseAfterVP_Fourth + output.LapseRiskOnExpAfterVP_Fourth)

    elif input.table_code_4 == "PURE":
        output.ProfitMarginPremAfterVP_Fourth = tool_workbook.cells(4, 383).value
        output.CleanRiskPremiumAfterVP_Fourth = tool_workbook.cells(4, 384).value
        output.initFixedPFeeAfterVP_Fourth = tool_workbook.cells(4, 385).value
        output.renewalFixedPFeeAfterVP_Fourth = tool_workbook.cells(4, 386).value
        output.claimsExpPremiumAfterVP_Fourth = tool_workbook.cells(4, 387).value
        output.renewalVariPFeeAfterVP_Fourth = tool_workbook.cells(4, 388).value
        output.DisribChannelSalesAfterVP_Fourth = tool_workbook.cells(4, 389).value
        output.InterAWSaleAfterVP_Fourth = tool_workbook.cells(4, 390).value
        output.InterUFSaleAfterVP_Fourth = tool_workbook.cells(4, 391).value
        output.InterUFLapseAfterVP_Fourth = tool_workbook.cells(4, 392).value
        output.LapseRiskOnExpAfterVP_Fourth = tool_workbook.cells(4, 393).value

        output.pure_o_tot_elements_after_VP = mathtool.trunc(
                                                        output.ProfitMarginPremAfterVP_Fourth + output.CleanRiskPremiumAfterVP_Fourth + output.initFixedPFeeAfterVP_Fourth +
                                                        output.renewalFixedPFeeAfterVP_Fourth + output.claimsExpPremiumAfterVP_Fourth + output.renewalVariPFeeAfterVP_Fourth +
                                                        output.DisribChannelSalesAfterVP_Fourth + output.InterAWSaleAfterVP_Fourth + output.InterUFSaleAfterVP_Fourth +
                                                        output.InterUFLapseAfterVP_Fourth + output.LapseRiskOnExpAfterVP_Fourth)

    elif input.table_code_4 == "PURO":
        output.ProfitMarginPremAfterVP_Fourth = tool_workbook.cells(4, 383).value
        output.CleanRiskPremiumAfterVP_Fourth = tool_workbook.cells(4, 384).value
        output.initFixedPFeeAfterVP_Fourth = tool_workbook.cells(4, 385).value
        output.renewalFixedPFeeAfterVP_Fourth = tool_workbook.cells(4, 386).value
        output.claimsExpPremiumAfterVP_Fourth = tool_workbook.cells(4, 387).value
        output.renewalVariPFeeAfterVP_Fourth = tool_workbook.cells(4, 388).value
        output.DisribChannelSalesAfterVP_Fourth = tool_workbook.cells(4, 389).value
        output.InterAWSaleAfterVP_Fourth = tool_workbook.cells(4, 390).value
        output.InterUFSaleAfterVP_Fourth = tool_workbook.cells(4, 391).value
        output.InterUFLapseAfterVP_Fourth = tool_workbook.cells(4, 392).value
        output.LapseRiskOnExpAfterVP_Fourth = tool_workbook.cells(4, 393).value

        output.pure_o_tot_elements_after_VP = mathtool.trunc(
                                                output.ProfitMarginPremAfterVP_Fourth + output.CleanRiskPremiumAfterVP_Fourth + output.initFixedPFeeAfterVP_Fourth +
                                                output.renewalFixedPFeeAfterVP_Fourth + output.claimsExpPremiumAfterVP_Fourth + output.renewalVariPFeeAfterVP_Fourth +
                                                output.DisribChannelSalesAfterVP_Fourth + output.InterAWSaleAfterVP_Fourth + output.InterUFSaleAfterVP_Fourth +
                                                output.InterUFLapseAfterVP_Fourth + output.LapseRiskOnExpAfterVP_Fourth)

    else:
        output.ProfitMarginPremAfterVP_Fourth = 0
        output.CleanRiskPremiumAfterVP_Fourth = 0
        output.initFixedPFeeAfterVP_Fourth = 0
        output.renewalFixedPFeeAfterVP_Fourth = 0
        output.claimsExpPremiumAfterVP_Fourth = 0
        output.renewalVariPFeeAfterVP_Fourth = 0
        output.DisribChannelSalesAfterVP_Fourth = 0
        output.InterAWSaleAfterVP_Fourth = 0
        output.InterUFSaleAfterVP_Fourth = 0
        output.InterUFLapseAfterVP_Fourth = 0
        output.LapseRiskOnExpAfterVP_Fourth = 0

    xl_calc.app.calculation = 'manual'

    # Setting total benefit premiums before and after VP
    xl_write.cells(x, 33).value = output.main_tot_before_VP
    xl_write.cells(x, 179).value = output.main_tot_after_VP
    xl_write.cells(x, 71).value = output.second_tot_before_VP
    xl_write.cells(x, 205).value = output.second_tot_after_VP
    xl_write.cells(x, 109).value = output.third_tot_before_VP
    xl_write.cells(x, 231).value = output.third_tot_after_VP
    xl_write.cells(x, 147).value = output.fourth_tot_before_VP
    xl_write.cells(x, 257).value = output.fourth_tot_after_VP

    # Setting elements before and after VP for main benefit
    xl_write.cells(x, 36).value = output.initFixedPFeeBeforeVP_Main
    xl_write.cells(x, 37).value = output.renewalFixedPFeeBeforeVP_Main
    xl_write.cells(x, 39).value = output.renewalVariPFeeBeforeVP_Main
    xl_write.cells(x, 40).value = output.DisribChannelSalesBeforeVP_Main
    xl_write.cells(x, 41).value = output.InterAWSaleBeforeVP_Main
    xl_write.cells(x, 42).value = output.InterUFSaleBeforeVP_Main
    xl_write.cells(x, 43).value = output.InterUFLapseBeforeVP_Main
    xl_write.cells(x, 44).value = output.CleanRiskPremiumBeforeVP_Main
    xl_write.cells(x, 45).value = output.LapseRiskOnExpBeforeVP_Main
    xl_write.cells(x, 46).value = output.ProfitMarginPremBeforeVP_Main
    xl_write.cells(x, 47).value = output.claimsExpPremiumBeforeVP_Main

    xl_write.cells(x, 181).value = output.initFixedPFeeAfterVP_Main
    xl_write.cells(x, 182).value = output.renewalFixedPFeeAfterVP_Main
    xl_write.cells(x, 184).value = output.renewalVariPFeeAfterVP_Main
    xl_write.cells(x, 185).value = output.DisribChannelSalesAfterVP_Main
    xl_write.cells(x, 186).value = output.InterAWSaleAfterVP_Main
    xl_write.cells(x, 187).value = output.InterUFSaleAfterVP_Main
    xl_write.cells(x, 188).value = output.InterUFLapseAfterVP_Main
    xl_write.cells(x, 189).value = output.CleanRiskPremiumAfterVP_Main
    xl_write.cells(x, 190).value = output.LapseRiskOnExpAfterVP_Main
    xl_write.cells(x, 191).value = output.ProfitMarginPremAfterVP_Main
    xl_write.cells(x, 192).value = output.claimsExpPremiumAfterVP_Main

    # Setting elements before and after VP for second benefit
    xl_write.cells(x, 74).value = output.initFixedPFeeBeforeVP_Second
    xl_write.cells(x, 75).value = output.renewalFixedPFeeBeforeVP_Second
    xl_write.cells(x, 77).value = output.renewalVariPFeeBeforeVP_Second
    xl_write.cells(x, 78).value = output.DisribChannelSalesBeforeVP_Second
    xl_write.cells(x, 79).value = output.InterAWSaleBeforeVP_Second
    xl_write.cells(x, 80).value = output.InterUFSaleBeforeVP_Second
    xl_write.cells(x, 81).value = output.InterUFLapseBeforeVP_Second
    xl_write.cells(x, 82).value = output.CleanRiskPremiumBeforeVP_Second
    xl_write.cells(x, 83).value = output.LapseRiskOnExpBeforeVP_Second
    xl_write.cells(x, 84).value = output.ProfitMarginPremBeforeVP_Second
    xl_write.cells(x, 85).value = output.claimsExpPremiumBeforeVP_Second

    xl_write.cells(x, 207).value = output.initFixedPFeeAfterVP_Second
    xl_write.cells(x, 208).value = output.renewalFixedPFeeAfterVP_Second
    xl_write.cells(x, 210).value = output.renewalVariPFeeAfterVP_Second
    xl_write.cells(x, 211).value = output.DisribChannelSalesAfterVP_Second
    xl_write.cells(x, 212).value = output.InterAWSaleAfterVP_Second
    xl_write.cells(x, 213).value = output.InterUFSaleAfterVP_Second
    xl_write.cells(x, 214).value = output.InterUFLapseAfterVP_Second
    xl_write.cells(x, 215).value = output.CleanRiskPremiumAfterVP_Second
    xl_write.cells(x, 216).value = output.LapseRiskOnExpAfterVP_Second
    xl_write.cells(x, 217).value = output.ProfitMarginPremAfterVP_Second
    xl_write.cells(x, 218).value = output.claimsExpPremiumAfterVP_Second

    # Setting elements before and after VP for Third benefit
    xl_write.cells(x, 112).value = output.initFixedPFeeBeforeVP_Third
    xl_write.cells(x, 113).value = output.renewalFixedPFeeBeforeVP_Third
    xl_write.cells(x, 115).value = output.renewalVariPFeeBeforeVP_Third
    xl_write.cells(x, 116).value = output.DisribChannelSalesBeforeVP_Third
    xl_write.cells(x, 117).value = output.InterAWSaleBeforeVP_Third
    xl_write.cells(x, 118).value = output.InterUFSaleBeforeVP_Third
    xl_write.cells(x, 119).value = output.InterUFLapseBeforeVP_Third
    xl_write.cells(x, 120).value = output.CleanRiskPremiumBeforeVP_Third
    xl_write.cells(x, 121).value = output.LapseRiskOnExpBeforeVP_Third
    xl_write.cells(x, 122).value = output.ProfitMarginPremBeforeVP_Third
    xl_write.cells(x, 123).value = output.claimsExpPremiumBeforeVP_Third

    xl_write.cells(x, 233).value = output.initFixedPFeeAfterVP_Third
    xl_write.cells(x, 234).value = output.renewalFixedPFeeAfterVP_Third
    xl_write.cells(x, 236).value = output.renewalVariPFeeAfterVP_Third
    xl_write.cells(x, 237).value = output.DisribChannelSalesAfterVP_Third
    xl_write.cells(x, 238).value = output.InterAWSaleAfterVP_Third
    xl_write.cells(x, 239).value = output.InterUFSaleAfterVP_Third
    xl_write.cells(x, 240).value = output.InterUFLapseAfterVP_Third
    xl_write.cells(x, 241).value = output.CleanRiskPremiumAfterVP_Third
    xl_write.cells(x, 242).value = output.LapseRiskOnExpAfterVP_Third
    xl_write.cells(x, 243).value = output.ProfitMarginPremAfterVP_Third
    xl_write.cells(x, 244).value = output.claimsExpPremiumAfterVP_Third

    # Setting elements before and after VP for fourth benefit
    xl_write.cells(x, 150).value = output.initFixedPFeeBeforeVP_Fourth
    xl_write.cells(x, 151).value = output.renewalFixedPFeeBeforeVP_Fourth
    xl_write.cells(x, 153).value = output.renewalVariPFeeBeforeVP_Fourth
    xl_write.cells(x, 154).value = output.DisribChannelSalesBeforeVP_Fourth
    xl_write.cells(x, 155).value = output.InterAWSaleBeforeVP_Fourth
    xl_write.cells(x, 156).value = output.InterUFSaleBeforeVP_Fourth
    xl_write.cells(x, 157).value = output.InterUFLapseBeforeVP_Fourth
    xl_write.cells(x, 158).value = output.CleanRiskPremiumBeforeVP_Fourth
    xl_write.cells(x, 159).value = output.LapseRiskOnExpBeforeVP_Fourth
    xl_write.cells(x, 160).value = output.ProfitMarginPremBeforeVP_Fourth
    xl_write.cells(x, 161).value = output.claimsExpPremiumBeforeVP_Fourth

    xl_write.cells(x, 259).value = output.initFixedPFeeAfterVP_Fourth
    xl_write.cells(x, 260).value = output.renewalFixedPFeeAfterVP_Fourth
    xl_write.cells(x, 262).value = output.renewalVariPFeeAfterVP_Fourth
    xl_write.cells(x, 263).value = output.DisribChannelSalesAfterVP_Fourth
    xl_write.cells(x, 264).value = output.InterAWSaleAfterVP_Fourth
    xl_write.cells(x, 265).value = output.InterUFSaleAfterVP_Fourth
    xl_write.cells(x, 266).value = output.InterUFLapseAfterVP_Fourth
    xl_write.cells(x, 267).value = output.CleanRiskPremiumAfterVP_Fourth
    xl_write.cells(x, 268).value = output.LapseRiskOnExpAfterVP_Fourth
    xl_write.cells(x, 269).value = output.ProfitMarginPremAfterVP_Fourth
    xl_write.cells(x, 270).value = output.claimsExpPremiumAfterVP_Fourth

    xl_calc.app.calculation = 'manual'

xl_write.book.save()
tool_workbook.book.close()
xl_write.book.close()

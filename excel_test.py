import openpyxl, openpyxl_templates, openpyxl_utilities
from openpyxl.styles import Color, PatternFill


scenario_source = "C:\\Projects\\Internal testing\\riskinternaltesting\\RiskInternalTesting-WEB\\src\\test\\resources\\7005135.xlsx"
output_file = "C:\\Users\\EcBerry\\Desktop\\book1.xlsx"
myworkbook = openpyxl.load_workbook(scenario_source)




        if value1 == value2:
            myworksheet.cell(row=7, column=1).fill = green_fill
        else:
            myworksheet.cell(row=7, column=1).fill = red_fill



class valuesMatch(object, val_1, val_2, workbook):

    object.value1 = None
    object.value2 = None
    object.workbook = None

    workbook = workbook
    myworksheet = myworkbook.active

    my_green = openpyxl.styles.colors.Color(rgb='00FF00')
    green_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_green)
    my_red = openpyxl.styles.colors.Color(rgb='FF0000')
    red_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)

    def __init__(self, object):
        self.object = object


    def __call__(self, *args, **kwargs):
        self.object()




values1 = valuesMatch(10, 10, myworkbook)    # object one
values2 = valuesMatch(10, 11, myworkbook)    # object two



@valuesMatch
def match1():
    #myworksheet = myworkbook.active
    #myworksheet.cell(row=7, column=2)
    #myworksheet.cell(row=7, column=2)
    print("Hello")



myworkbook.save(output_file)
